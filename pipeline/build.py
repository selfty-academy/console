#!/usr/bin/env python3
"""Console Selfty Academy : xlsx « Candidature Webi » -> console.html.

Usage : python3 build.py [chemin/vers/candidature-webi.xlsx]
Le xlsx = export du Sheet « Candidature Webi »
(1mKA765MImL3103Foil5kYu1IR14Ea55nOIv4n7UbHCc), onglets Inscriptions,
Visites, « Mail a contacter webi  ».
"""
import sys, re, json, base64, datetime, io, os
from collections import Counter, OrderedDict
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else HERE / "candidature-webi.xlsx"
TEST_EMAILS = {"alexyoucompte99@gmail.com", "alexandre.majorel@tsm-education.fr",
               "anaisbrault86@gmail.com"}  # Anaïs teste son propre funnel
# SHOW_TEST=1 (build local uniquement, jamais en CI) : garde les e-mails de test dans Clientes/Contrats
# et ajoute les faux calls de test-calls.json (gitignoré) pour tester la console de bout en bout
SHOW_TEST = os.environ.get("SHOW_TEST") == "1"

def norm_phone(v):
    if v is None:
        return ""
    if isinstance(v, float):
        v = int(v)
    s = str(v).strip()
    if not s or s.startswith("#"):
        return ""
    had_plus = s.startswith("+")
    d = re.sub(r"\D", "", s)
    if not d:
        return ""
    if had_plus:
        return d
    if d.startswith("00"):
        return d[2:]
    if d.startswith("0") and len(d) == 10:
        return "33" + d[1:]
    if len(d) == 9 and d[0] in "67":
        return "33" + d
    return d

def fmt_date(dt):
    return dt.strftime("%d/%m %H:%M") if isinstance(dt, datetime.datetime) else ""

def src_label(s):
    s = (s or "").lower()
    if "utm_source=email" in s or "utm_medium=email" in s:
        return "E-mail"
    if "link_in_bio" in s:
        return "Bio Insta"
    if "story" in s:
        return "Story Insta"
    if "utm_source=ig" in s or "instagram" in s:
        return "Insta autre"
    return "Direct / autre"

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Inscriptions"]
hdr = [c.value for c in ws[1]]
rows = [dict(zip(hdr, [c.value for c in r])) for r in ws.iter_rows(min_row=2) if any(c.value for c in r)]
real = [r for r in rows
        if (str(r.get("E-mail") or "").strip().lower() not in TEST_EMAILS)
        and not str(r.get("Prénom") or "").lower().startswith("alex test")]

inscrits, cands, la = [], [], []
for r in real:
    tel = norm_phone(r.get("Téléphone"))
    base = {
        "n": str(r.get("Prénom") or "?").strip(),
        "tel": tel,
        "mail": str(r.get("E-mail") or "").strip(),
        "mode": r.get("Mode") or "",
        "date": fmt_date(r.get("Date")),
        "ts": r.get("Date").isoformat() if isinstance(r.get("Date"), datetime.datetime) else "",
        "src": src_label(r.get("Source")),
        "statut": str(r.get("Statut ") or "").strip(),
        "etape": r.get("Dernière étape") or "",
    }
    det = []
    for col, lab in [("Situation", "Sa situation"), ("Sujet du coaching", "Sujet du coaching"),
                     ("Déjà essayé", "Déjà essayé"), ("Accord coaching live", "Accord coaching en direct"),
                     ("LA1 · Parcours", "Parcours"), ("LA2 · Résultat idéal", "Résultat idéal"),
                     ("LA3 · Envie d'apprendre", "Envie d'apprendre"), ("LA4 · Différence", "Ce qui ferait la différence")]:
        v = str(r.get(col) or "").strip()
        if v and v.lower() != "non":
            det.append([lab, v])
    base["det"] = det
    inscrits.append(base)
    if r.get("Mode") == "coaching":
        cands.append({**base,
                      "situation": str(r.get("Situation") or "").strip(),
                      "sujet": str(r.get("Sujet du coaching") or "").strip(),
                      "deja": str(r.get("Déjà essayé") or "").strip(),
                      "accord": str(r.get("Accord coaching live") or "").strip()})
    if str(r.get("Liste d'attente") or "").strip().lower() == "oui":
        la.append({**base,
                   "la1": str(r.get("LA1 · Parcours") or "").strip(),
                   "la2": str(r.get("LA2 · Résultat idéal") or "").strip(),
                   "la3": str(r.get("LA3 · Envie d'apprendre") or "").strip(),
                   "la4": str(r.get("LA4 · Différence") or "").strip()})

la_deja = sum(1 for r in real if r.get("Dernière étape") == "liste_attente_deja_inscrite")

# ---- Liste d'attente école (Sheet « École de coaching  (réponses) ») ----
ewb = openpyxl.load_workbook(HERE / "liste-attente.xlsx", data_only=True)
ews = ewb.active
ehdr = [str(c.value or "").strip() for c in ews[1]]
def short_label(h):
    return h.split("\n")[0].strip().rstrip('?" ').strip() or h[:40]
FORM_COLS = [
    ("Es-tu déjà coach", "Parcours"),
    ("Quel serait ton résultat", "Résultat idéal"),
    ("Qu'est ce que tu as le plus envie d'apprendre", "Envie d'apprendre"),
    ("Qu'est ce qui selon toi ferait la différence", "Ce qui ferait la différence"),
]
ecole = []
for r in ews.iter_rows(min_row=2):
    row = dict(zip(ehdr, [c.value for c in r]))
    if not any(row.values()):
        continue
    nom = str(row.get("Nom prénom et age") or "").strip()
    mail = str(row.get("ton e-mail") or "").strip().lower()
    if not nom and not mail:
        continue
    detail = []
    for pref, lab in FORM_COLS:
        for h in ehdr:
            if h.startswith(pref) and row.get(h):
                detail.append([lab, str(row[h]).strip()])
                break
    notes = []
    idx_start = ehdr.index("ton e-mail") + 5 if "ton e-mail" in ehdr else 12
    for h in ehdr[idx_start:]:
        v = row.get(h)
        if v and str(v).strip() and h not in ("Enregistrement appel",):
            notes.append([short_label(h)[:60], str(v).strip()])
    dt = row.get("Horodateur")
    ecole.append({
        "n": nom or mail,
        "mail": mail,
        "tel": norm_phone(row.get("Numéro de téléphone")),
        "date": fmt_date(dt),
        "ts": dt.isoformat() if isinstance(dt, datetime.datetime) else "",
        "statut": str(row.get("Statut") or "").strip(),
        "chaud": str(row.get("Chaud pour closing ?") or "").strip(),
        "qui": str(row.get("Qui prend ?") or "").strip(),
        "comm": str(row.get("Commentaire") or "").strip(),
        "rec": str(row.get("Enregistrement appel") or "").strip(),
        "detail": detail,
        "notes": notes,
        "src": "Formulaire école",
    })
# + candidatures liste d'attente venues de la page du live, absentes du Sheet
ecole_mails = {e["mail"] for e in ecole if e["mail"]}
for c in la:
    if c["mail"].lower() not in ecole_mails:
        ecole.append({
            "n": c["n"], "mail": c["mail"], "tel": c["tel"], "date": c["date"],
            "ts": c["ts"], "statut": "", "chaud": "", "qui": "", "comm": "",
            "rec": "", "src": "Page du live",
            "detail": [x for x in [["Parcours", c["la1"]], ["Résultat idéal", c["la2"]],
                                   ["Envie d'apprendre", c["la3"]], ["Ce qui ferait la différence", c["la4"]]] if x[1]],
            "notes": [],
        })
ecole.sort(key=lambda e: e["ts"], reverse=True)
ecole_uniques = len({e["mail"] or e["n"] for e in ecole if "doublon" not in e["statut"].lower()})
# recoupement : inscrites au webi (n'importe quel mode)
webi_mails = {i["mail"].lower() for i in inscrits if i["mail"]}
for e in ecole:
    e["webi"] = bool(e["mail"] and e["mail"] in webi_mails)
# segments des inscrits webi : candidat coaching / intéressé école
eco_mails = {e["mail"] for e in ecole if e["mail"]}
ecole_by_mail_all = {e["mail"]: e for e in reversed(ecole) if e["mail"]}
for i in inscrits:
    i["coach"] = i["mode"] == "coaching"
    i["eco"] = bool(i["mail"] and i["mail"].lower() in eco_mails)
    if i["eco"]:
        fiche = ecole_by_mail_all.get(i["mail"].lower())
        if fiche:
            deja = {d[0] for d in i["det"]}
            i["det"] += [d for d in fiche["detail"] if d[0] not in deja]

# ---- Compta : onglets « Paiements » et « Charges » du Sheet École (optionnels) ----
def read_tab(wb_, name):
    for ws_ in wb_.worksheets:
        if ws_.title.strip().lower() == name:
            h = [str(c.value or "").strip() for c in ws_[1]]
            return [dict(zip(h, [c.value for c in r])) for r in ws_.iter_rows(min_row=2) if any(c.value for c in r)]
    return None

def eur(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d,.\-]", "", str(v)).replace(",", ".")
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0

paiements_rows = read_tab(ewb, "paiements")
charges_rows = read_tab(ewb, "charges")
compta_ok = paiements_rows is not None or charges_rows is not None
paiements, charges = [], []
for r in (paiements_rows or []):
    if not (r.get("Client") or r.get("E-mail")):
        continue
    dt = r.get("Date")
    paiements.append({
        "date": fmt_date(dt) if isinstance(dt, datetime.datetime) else str(dt or "").strip(),
        "ts": dt.isoformat() if isinstance(dt, datetime.datetime) else "",
        "client": str(r.get("Client") or "").strip(),
        "mail": str(r.get("E-mail") or "").strip().lower(),
        "montant": eur(r.get("Montant")),
        "total": eur(r.get("Prix total")),
        "note": str(r.get("Note") or "").strip(),
    })
for r in (charges_rows or []):
    if not (r.get("Poste") or r.get("Montant")):
        continue
    dt = r.get("Date")
    charges.append({
        "date": fmt_date(dt) if isinstance(dt, datetime.datetime) else str(dt or "").strip(),
        "ts": dt.isoformat() if isinstance(dt, datetime.datetime) else "",
        "poste": str(r.get("Poste") or "?").strip(),
        "montant": eur(r.get("Montant")),
        "note": str(r.get("Note") or "").strip(),
    })
# ---- Suivi calls (onglet « Suivi Calls », écrit par la console via le pont) ----
track_rows = read_tab(ewb, "suivi calls")
track = {}
for r in (track_rows or []):
    cid = str(r.get("Call ID") or "").strip()
    if not cid or (cid.upper().startswith("TEST") and not SHOW_TEST):
        continue
    track[cid] = {
        "s": str(r.get("Show-up") or "").strip(),
        "r": str(r.get("Résultat") or "").strip(),
        "prix": eur(r.get("Prix")),
        "np": int(eur(r.get("Nb paiements")) or 0),
        "q": int(eur(r.get("Qualif /10")) or 0),
        "retrans": str(r.get("Retranscription") or "").strip(),
        "comment": str(r.get("Commentaire") or "").strip(),
    }

# ---- Clientes signées (onglet « Clients », créé par le pont à la 1re vente) ----
clients_rows = read_tab(ewb, "clients")
clientes = []
for r in (clients_rows or []):
    mail = str(r.get("E-mail") or "").strip().lower()
    nom = str(r.get("Nom") or "").strip()
    if (not mail and not nom) or (mail in TEST_EMAILS and not SHOW_TEST):
        continue
    dt = r.get("Date signature")
    clientes.append({
        "n": nom or mail,
        "mail": mail,
        "tel": norm_phone(r.get("Téléphone")),
        "date": dt.strftime("%d/%m/%Y") if isinstance(dt, datetime.datetime) else str(dt or "").strip(),
        "ts": dt.isoformat() if isinstance(dt, datetime.datetime) else "",
        "offre": str(r.get("Offre") or "").strip(),
        "prix": eur(r.get("Prix")),
        "statut": str(r.get("Statut") or "").strip(),
        "notes": str(r.get("Notes") or "").strip(),
    })
clientes.sort(key=lambda c: c["ts"], reverse=True)

# ---- Contrats envoyés / signés (onglet « Contrats », écrit par le pont) ----
def dstr(v):
    return v.strftime("%d/%m/%Y %H:%M") if isinstance(v, datetime.datetime) else str(v or "").strip()

contrats_rows = read_tab(ewb, "contrats")
contrats = {}
for r in (contrats_rows or []):
    mail = str(r.get("E-mail") or "").strip().lower()
    if not mail or (mail in TEST_EMAILS and not SHOW_TEST):
        continue
    st = str(r.get("Statut") or "").strip()
    rec = {
        "token": str(r.get("Token") or "").strip(),
        "mail": mail,
        "prenom": str(r.get("Prénom") or "").strip(),
        "nom": str(r.get("Nom") or "").strip(),
        "prix": eur(r.get("Prix")),
        "np": int(eur(r.get("Nb paiements")) or 1),
        "statut": st,
        "sent": dstr(r.get("Date envoi")),
        "signedAt": dstr(r.get("Date signature")),
        "pdf": str(r.get("PDF") or "").strip(),
        "caseA": str(r.get("Case A") or "").strip() == "Oui",
        "caseB": str(r.get("Case B") or "").strip() == "Oui",
    }
    prev = contrats.get(mail)
    # un contrat signé prime ; sinon le plus récent (dernière ligne) l'emporte
    if not prev or st == "Signé" or prev["statut"] != "Signé":
        contrats[mail] = rec

# lien automatique paiements -> suivi des appels (fiche école par e-mail)
ecole_by_mail = {e["mail"]: e for e in ecole if e["mail"]}
clients = OrderedDict()
for p in sorted(paiements, key=lambda x: x["ts"]):
    k = p["mail"] or p["client"].lower()
    c = clients.setdefault(k, {"nom": p["client"] or p["mail"], "mail": p["mail"],
                              "paiements": [], "recu": 0.0, "total": 0.0})
    c["paiements"].append({"date": p["date"], "montant": p["montant"], "note": p["note"]})
    c["recu"] += p["montant"]
    c["total"] = max(c["total"], p["total"])
    suivi = ecole_by_mail.get(p["mail"])
    if suivi:
        c["appel"] = {"qui": suivi["qui"], "statut": suivi["statut"], "chaud": suivi["chaud"]}
# paiements -> fiches clientes (encaissé / contracté)
for c in clientes:
    p = clients.get(c["mail"])
    c["recu"] = p["recu"] if p else 0.0
    c["total"] = max(c["prix"], p["total"] if p else 0.0, c["recu"])

# Visites
vws = wb["Visites"]
visites = [[c.value for c in r] for r in vws.iter_rows(min_row=2) if any(c.value for c in r)]
v_mobile = sum(1 for v in visites if str(v[1]).strip().lower() == "mobile")

# Campagne mail
mws = wb["Mail a contacter webi "]
mhdr = [str(c.value or "").strip() for c in mws[1]]
i_statut = mhdr.index("Statut envoi") if "Statut envoi" in mhdr else None
mails_total, mails_envoyes = 0, 0
for r in mws.iter_rows(min_row=2):
    vals = [c.value for c in r]
    if not any(vals):
        continue
    mails_total += 1
    if i_statut is not None and vals[i_statut] and "envoyé" in str(vals[i_statut]).lower():
        mails_envoyes += 1

# Inscriptions par jour
par_jour = OrderedDict()
for i in sorted(inscrits, key=lambda x: x["ts"]):
    if i["ts"]:
        d = i["ts"][5:10]
        key = d[3:5] + "/" + d[0:2]
        par_jour[key] = par_jour.get(key, 0) + 1

sources = Counter(i["src"] for i in inscrits)

# ---- iClosed : tous les calls bookés (passés, à venir, annulés) ----
import os
import urllib.request

ic_key = os.environ.get("ICLOSED_KEY", "")
if not ic_key and (HERE / "iclosed-key.txt").exists():
    ic_key = (HERE / "iclosed-key.txt").read_text().strip()
icalls, ic_ok = [], False
if ic_key:
    try:
        raw = []
        for page in range(0, 20):  # pages 0-indexées, 100 par page, garde-fou 2000 calls
            req = urllib.request.Request(
                f"https://public.api.iclosed.io/v1/eventCalls?limit=100&page={page}",
                headers={"Authorization": "Bearer " + ic_key})
            batch = json.load(urllib.request.urlopen(req, timeout=30)).get("data", {}).get("eventCalls", [])
            raw += batch
            if len(batch) < 100:
                break
        seen_ids = set()
        for c in raw:
            if c.get("id") in seen_ids:
                continue
            seen_ids.add(c.get("id"))
            quest = []
            for q in c.get("secondaryAnswers") or []:
                ans = " / ".join(str(a.get("answer") or "") for a in (q.get("answer") or []) if a.get("answer"))
                if ans:
                    quest.append([str(q.get("statement") or "?").strip(), ans])
            task = (c.get("task") or [{}])[0]
            icalls.append({
                "id": c.get("id"),
                "n": str(c.get("inviteeName") or "?").strip(),
                "mail": str(c.get("inviteeEmail") or "").strip().lower(),
                "tel": norm_phone(c.get("phoneNumber")),
                "utc": c.get("dateTimeUTC") or "",
                "link": c.get("locationLinkInvitee") or "",
                "event": str((c.get("event") or {}).get("name") or "").strip(),
                "closer": str((c.get("user") or {}).get("firstName") or "").strip(),
                "cancel": bool(c.get("cancelReason")) or c.get("eventType") == "CANCELLED",
                "cancelWhy": str(c.get("cancelReason") or "").strip(),
                "outcome": str(task.get("outcome") or "").strip(),
                "notes": str(task.get("notes") or c.get("notes") or "").strip(),
                "quest": quest,
            })
        ic_ok = True
        print(f"iClosed : {len(icalls)} calls")
    except Exception as ex:
        print("iClosed fetch KO (on garde la console sans) :", ex)
# faux calls de test : env TEST_CALLS (secret GitHub, pour le CI) ou fichier local test-calls.json
if SHOW_TEST:
    raw_tc = os.environ.get("TEST_CALLS", "") or ((HERE / "test-calls.json").read_text() if (HERE / "test-calls.json").exists() else "")
    if raw_tc.strip():
        icalls += json.loads(raw_tc)
        print("SHOW_TEST : faux calls ajoutés")
# suivi closing du Sheet accroché à chaque call ; « Call test » = exclu de partout
for c in icalls:
    c["trk"] = track.get(str(c["id"]))
icalls = [c for c in icalls if not (c["trk"] and c["trk"]["s"].lower() == "call test")]
if ic_ok:
    # dump minimal pour notify_calls.py (notif Telegram des nouveaux bookings)
    (HERE / "icalls.json").write_text(json.dumps(
        [{"id": c["id"], "n": c["n"], "utc": c["utc"], "event": c["event"], "cancel": c["cancel"]}
         for c in icalls], ensure_ascii=False))

# ---- Scholarship : candidatures Tally (form Np1Gy0, compte perso Alex) ----
try:
    from zoneinfo import ZoneInfo
    TZ_PARIS = ZoneInfo("Europe/Paris")
except Exception:
    TZ_PARIS = None

def iso_paris(at):
    try:
        d = datetime.datetime.fromisoformat(at.replace("Z", "+00:00"))
        if TZ_PARIS:
            d = d.astimezone(TZ_PARIS)
        return d.strftime("%d/%m %H:%M")
    except Exception:
        return ""

ty_key = os.environ.get("TALLY_API_KEY", "")
if not ty_key and (HERE / "tally-key.txt").exists():
    ty_key = (HERE / "tally-key.txt").read_text().strip()
schol_subs, schol_ok, schol_stats = [], False, {}
SCHOL_SKIP = TEST_EMAILS | {"test@test.fr"}
ID_PRENOM, ID_NOM, ID_MAIL, ID_TEL, ID_INSTA, ID_HIDDEN = "yE0EXd", "XMRM5z", "8PJPNr", "0JbJVA", "zr0rEg", "g7Q7bJ"

def ty_txt(a):
    if a is None:
        return ""
    if isinstance(a, list):
        return " · ".join(str(x) for x in a if not isinstance(x, dict))
    return str(a).strip()

if ty_key:
    try:
        qlabels, raw_subs, page = {}, [], 1
        while True:
            req = urllib.request.Request(
                f"https://api.tally.so/forms/Np1Gy0/submissions?filter=all&page={page}",
                headers={"Authorization": "Bearer " + ty_key,
                         "User-Agent": "curl/8.4.0"})  # Cloudflare bloque l'UA Python
            d = json.load(urllib.request.urlopen(req, timeout=30))
            for q in d.get("questions") or []:
                qlabels[q["id"]] = (str(q.get("title") or "?").strip(), str(q.get("type") or ""))
            raw_subs += d.get("submissions") or []
            schol_stats = d.get("totalNumberOfSubmissionsPerFilter") or {}
            if not d.get("hasMore"):
                break
            page += 1
        for s in raw_subs:
            a = {r.get("questionId"): r.get("answer") for r in s.get("responses") or []}
            mail = ty_txt(a.get(ID_MAIL)).lower()
            if mail in SCHOL_SKIP:
                continue
            # personne n'a rien rempli d'identifiable : du bruit, on saute
            if not mail and not ty_txt(a.get(ID_PRENOM)) and not ty_txt(a.get(ID_TEL)):
                continue
            det, files = [], []
            for r in s.get("responses") or []:
                qid = r.get("questionId")
                if qid in (ID_PRENOM, ID_NOM, ID_MAIL, ID_TEL, ID_INSTA, ID_HIDDEN):
                    continue
                lab, qtype = qlabels.get(qid, ("?", ""))
                ans = r.get("answer")
                if qtype == "FILE_UPLOAD" and isinstance(ans, list):
                    files += [{"n": str(f.get("name") or "fichier"), "u": str(f.get("url") or "")}
                              for f in ans if isinstance(f, dict)]
                    continue
                v = ty_txt(ans)
                if v:
                    det.append([lab, v])
            hid = a.get(ID_HIDDEN) if isinstance(a.get(ID_HIDDEN), dict) else {}
            at = str(s.get("submittedAt") or "")
            schol_subs.append({
                "id": s.get("id"),
                "n": (ty_txt(a.get(ID_PRENOM)) + " " + ty_txt(a.get(ID_NOM))).strip() or mail or "?",
                "mail": mail,
                "tel": norm_phone(ty_txt(a.get(ID_TEL))),
                "insta": ty_txt(a.get(ID_INSTA)).lstrip("@"),
                "at": at,
                "date": iso_paris(at),
                "done": bool(s.get("isCompleted")),
                "src": str(hid.get("source") or "").strip(),
                "det": det,
                "files": files,
            })
        schol_subs.sort(key=lambda x: x["at"], reverse=True)
        schol_ok = True
        print(f"Scholarship Tally : {len(schol_subs)} candidature(s)")
    except Exception as ex:
        print("Tally scholarship KO (on garde la console sans) :", ex)


data = {
    "maj": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
    "webi": {
        "label": "Live du lundi 31 août, 18h",
        "meet": "https://meet.google.com/oxf-vzjg-bhr",
        "groupe": "https://chat.whatsapp.com/JRCRXWKVg8qBUS0uUUTkCA?mode=gi_t",
        "lp": "https://selfty-academy.github.io/live-31-aout/",
    },
    "inscrits": sorted(inscrits, key=lambda x: x["ts"], reverse=True),
    "cands": sorted(cands, key=lambda c: (c["accord"] != "oui", c["ts"])),
    "la": la,
    "laDeja": la_deja,
    "ecole": ecole,
    "ecoleUniques": ecole_uniques,
    "icalls": {"ok": ic_ok, "calls": sorted(icalls, key=lambda x: x["utc"])},
    "suivi": {
        "trackOk": track_rows is not None,
        "clientsOk": clients_rows is not None,
        "clientes": clientes,
        "contratsOk": contrats_rows is not None,
        "contrats": list(contrats.values()),
    },
    "schol": {"ok": schol_ok, "url": "https://tally.so/r/Np1Gy0",
              "stats": schol_stats, "subs": schol_subs},
    "compta": {
        "ok": compta_ok,
        "sheetUrl": "https://docs.google.com/spreadsheets/d/1CUiT962_dGEAWhydaboYmC23ir8gA-CtZyUXB4gErIc/edit",
        "clients": list(clients.values()),
        "paiements": paiements,
        "charges": sorted(charges, key=lambda x: x["ts"], reverse=True),
    },
    "stats": {
        "visites": len(visites),
        "vMobile": v_mobile,
        "inscrits": len(inscrits),
        "coaching": len(cands),
        "mailsTotal": mails_total,
        "mailsEnvoyes": mails_envoyes,
        "parJour": list(par_jour.items()),
        "sources": dict(sources.most_common()),
    },
}

# Logo en data URI (réduit)
try:
    from PIL import Image
    im = Image.open(HERE / "logo-selfty-encre.png")
    im.thumbnail((360, 360))
    buf = io.BytesIO()
    im.save(buf, "PNG", optimize=True)
    logo = "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()
except Exception:
    logo = "data:image/png;base64," + base64.b64encode((HERE / "logo-selfty-encre.png").read_bytes()).decode()

# Pont Apps Script (suivi école éditable) : pont.json local ou env (CI)
pont = {"url": "", "key": ""}
pont_file = HERE / "pont.json"
if pont_file.exists():
    pont.update(json.loads(pont_file.read_text()))
pont["url"] = os.environ.get("PONT_URL", pont["url"])
pont["key"] = os.environ.get("PONT_KEY", pont["key"])

tpl = (HERE / "template.html").read_text()
out = (tpl.replace("__DATA__", json.dumps(data, ensure_ascii=False)).replace("__LOGO__", logo)
       .replace("__PONT_URL__", pont["url"]).replace("__PONT_KEY__", pont["key"]))
(HERE / "console.html").write_text(out)
print(f"console.html : {len(inscrits)} inscrits, {len(cands)} candidatures live, {len(ecole)} lignes école ({ecole_uniques} personnes), {len(visites)} visites")
